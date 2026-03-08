# backend/app/services/excel_io.py
"""Core logic for Excel import/export of transactions."""
import re
import uuid
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from fastapi import HTTPException
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.transaction import Transaction
from app.models.user_card import UserCard
from app.schemas.excel_io import ColumnMapping, ImportConfirmResponse, ImportPreviewResponse
from app.services import import_cache
from app.services.category import list_categories
from app.services.transaction import list_transactions

# ── Header pattern matching ──────────────────────────────────────────────────

FIELD_PATTERNS: dict[str, list[str]] = {
    "transacted_at": ["날짜", "date", "일시", "거래일", "결제일", "사용일", "일자"],
    "amount": ["금액", "amount", "결제금액", "거래금액", "사용금액", "출금액", "입금액"],
    "description": ["내역", "description", "적요", "거래내역", "가맹점", "이용내역", "사용처", "비고"],
    "type": ["유형", "type", "거래유형", "구분", "입출금", "거래구분"],
    "category_name": ["카테고리", "category", "분류", "항목"],
    "payment_type": ["결제수단", "payment", "결제방법"],
    "card_name": ["카드", "card", "카드명"],
}

TYPE_MAP: dict[str, str] = {
    "지출": "expense",
    "출금": "expense",
    "수입": "income",
    "입금": "income",
    "이체": "transfer",
    "송금": "transfer",
    "expense": "expense",
    "income": "income",
    "transfer": "transfer",
}

# ── Auto-detect column mapping ───────────────────────────────────────────────


def auto_detect_mapping(headers: list[str]) -> ColumnMapping:
    """Match header names to field names using pattern matching."""
    mapping: dict[str, int | None] = {}
    lower_headers = [h.lower().strip() for h in headers]
    for field, patterns in FIELD_PATTERNS.items():
        for idx, header in enumerate(lower_headers):
            if any(p in header for p in patterns):
                mapping[field] = idx
                break
    return ColumnMapping(**mapping)


# ── Parse and preview ────────────────────────────────────────────────────────


def parse_and_preview(file_bytes: bytes) -> ImportPreviewResponse:
    """Parse an xlsx file and return a preview with auto-detected mapping."""
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="올바른 xlsx 파일이 아닙니다.")

    ws = wb.active
    if ws is None:
        wb.close()
        raise HTTPException(status_code=400, detail="빈 엑셀 파일입니다.")

    all_rows: list[list] = []
    for row in ws.iter_rows():
        all_rows.append([cell.value for cell in row])
    wb.close()

    if len(all_rows) == 0:
        raise HTTPException(status_code=400, detail="빈 엑셀 파일입니다.")

    headers = [str(v) if v is not None else "" for v in all_rows[0]]
    data_rows = all_rows[1:]

    if len(data_rows) == 0:
        raise HTTPException(status_code=400, detail="데이터 행이 없습니다.")

    auto_mapping = auto_detect_mapping(headers)

    # Convert all cell values to strings for preview
    preview_rows: list[list[str | None]] = []
    for row in data_rows[:10]:
        preview_rows.append([str(v) if v is not None else None for v in row])

    # Store in cache
    # Convert data rows for cache (keep original types for parsing later)
    cache_rows = []
    for row in data_rows:
        cache_rows.append([v for v in row])

    import_id = import_cache.store(headers, cache_rows)

    return ImportPreviewResponse(
        import_id=import_id,
        headers=headers,
        auto_mapping=auto_mapping,
        preview_rows=preview_rows,
        total_rows=len(data_rows),
    )


# ── Date parsing ─────────────────────────────────────────────────────────────

_DATE_PATTERNS = [
    (re.compile(r"^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})"), "%Y-%m-%d"),
    (re.compile(r"^(\d{2})[-/.](\d{1,2})[-/.](\d{1,2})"), "%y-%m-%d"),
]


def _parse_date(value) -> datetime | None:
    """Try to parse a date value into a datetime."""
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc)
        return value
    if not value:
        return None
    s = str(value).strip()
    for pattern, _ in _DATE_PATTERNS:
        m = pattern.match(s)
        if m:
            try:
                y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
                if y < 100:
                    y += 2000
                return datetime(y, mo, d, tzinfo=timezone.utc)
            except ValueError:
                continue
    return None


# ── Amount parsing ───────────────────────────────────────────────────────────


def _parse_amount(value) -> Decimal | None:
    """Parse an amount value, removing commas and currency symbols."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return Decimal(str(abs(value)))
    s = str(value).strip()
    s = re.sub(r"[원₩\s,]", "", s)
    if not s:
        return None
    try:
        return abs(Decimal(s))
    except InvalidOperation:
        return None


# ── Type parsing ─────────────────────────────────────────────────────────────


def _parse_type(value, default_type: str) -> str:
    """Parse transaction type from a cell value."""
    if value is None:
        return default_type
    s = str(value).strip().lower()
    return TYPE_MAP.get(s, default_type)


# ── Confirm import ───────────────────────────────────────────────────────────


def confirm_import(
    db: Session,
    user_id: uuid.UUID,
    import_id: str,
    mapping: ColumnMapping,
    default_type: str = "expense",
) -> ImportConfirmResponse:
    """Validate mapping, create transactions from cached data."""
    cached = import_cache.retrieve(import_id)
    if cached is None:
        raise HTTPException(status_code=404, detail="미리보기가 만료되었습니다. 다시 업로드해 주세요.")

    if mapping.transacted_at is None or mapping.amount is None:
        raise HTTPException(status_code=400, detail="날짜와 금액 컬럼은 필수입니다.")

    rows = cached["rows"]

    # Build name→id lookups
    categories = list_categories(db, user_id)
    cat_map = {c.name.lower(): c.id for c in categories}

    cards = list(db.scalars(select(UserCard).where(UserCard.user_id == user_id)).all())
    card_map = {c.name.lower(): c.id for c in cards}

    # Existing transaction dedup keys (normalize amount to int for comparison)
    existing_txs = list_transactions(db, user_id)
    dedup_keys: set[str] = set()
    for tx in existing_txs:
        date_str = tx.transacted_at.strftime("%Y-%m-%d") if tx.transacted_at else ""
        amt_normalized = str(int(tx.amount)) if tx.amount == int(tx.amount) else str(tx.amount)
        dedup_keys.add(f"{date_str}:{amt_normalized}:{tx.description or ''}")

    created_count = 0
    duplicate_count = 0
    error_count = 0
    errors: list[dict] = []
    new_transactions: list[Transaction] = []

    def _cell(row: list, idx: int | None):
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    for row_idx, row in enumerate(rows):
        try:
            # Parse date
            dt = _parse_date(_cell(row, mapping.transacted_at))
            if dt is None:
                errors.append({"row": row_idx + 2, "message": "날짜를 파싱할 수 없습니다."})
                error_count += 1
                continue

            # Parse amount
            amount = _parse_amount(_cell(row, mapping.amount))
            if amount is None or amount == 0:
                errors.append({"row": row_idx + 2, "message": "금액을 파싱할 수 없습니다."})
                error_count += 1
                continue

            # Parse other fields
            description = str(_cell(row, mapping.description)) if mapping.description is not None and _cell(row, mapping.description) is not None else None
            tx_type = _parse_type(_cell(row, mapping.type), default_type)
            payment_type = str(_cell(row, mapping.payment_type)).strip() if mapping.payment_type is not None and _cell(row, mapping.payment_type) is not None else None
            category_name = str(_cell(row, mapping.category_name)).strip().lower() if mapping.category_name is not None and _cell(row, mapping.category_name) is not None else None
            card_name_val = str(_cell(row, mapping.card_name)).strip().lower() if mapping.card_name is not None and _cell(row, mapping.card_name) is not None else None

            # Map payment_type to valid values
            payment_type_map = {
                "현금": "cash", "cash": "cash",
                "신용카드": "credit_card", "credit_card": "credit_card", "credit": "credit_card",
                "체크카드": "debit_card", "debit_card": "debit_card", "debit": "debit_card",
                "계좌이체": "bank", "bank": "bank", "이체": "bank",
            }
            if payment_type:
                payment_type = payment_type_map.get(payment_type.lower())

            # Lookup category/card IDs
            category_id = cat_map.get(category_name) if category_name else None
            user_card_id = card_map.get(card_name_val) if card_name_val else None

            # Dedup check
            date_str = dt.strftime("%Y-%m-%d")
            amt_normalized = str(int(amount)) if amount == int(amount) else str(amount)
            dedup_key = f"{date_str}:{amt_normalized}:{description or ''}"
            if dedup_key in dedup_keys:
                duplicate_count += 1
                continue

            dedup_keys.add(dedup_key)
            new_transactions.append(Transaction(
                user_id=user_id,
                category_id=category_id,
                type=tx_type,
                amount=amount,
                description=description,
                transacted_at=dt,
                payment_type=payment_type,
                user_card_id=user_card_id,
            ))
            created_count += 1
        except Exception as e:
            errors.append({"row": row_idx + 2, "message": str(e)})
            error_count += 1

    if new_transactions:
        db.add_all(new_transactions)
        db.commit()

    import_cache.remove(import_id)

    return ImportConfirmResponse(
        created_count=created_count,
        duplicate_count=duplicate_count,
        error_count=error_count,
        errors=errors,
    )


# ── Export ───────────────────────────────────────────────────────────────────


def export_transactions(
    db: Session,
    user_id: uuid.UUID,
    period: str = "month",
    year: int | None = None,
    month: int | None = None,
) -> BytesIO:
    """Export transactions to an xlsx file."""
    from datetime import date

    now = datetime.now(timezone.utc)
    if year is None:
        year = now.year
    if month is None:
        month = now.month

    # Calculate date range
    if period == "month":
        from_date = date(year, month, 1)
        if month == 12:
            to_date = date(year + 1, 1, 1)
        else:
            to_date = date(year, month + 1, 1)
        # to_date should be last day of month (list_transactions uses < next_day)
        from datetime import timedelta
        to_date = to_date - timedelta(days=1)
    elif period == "year":
        from_date = date(year, 1, 1)
        to_date = date(year, 12, 31)
    else:  # "all"
        from_date = None
        to_date = None

    transactions = list_transactions(db, user_id, from_date=from_date, to_date=to_date)

    # Build id→name lookups
    categories = list_categories(db, user_id)
    cat_id_map = {c.id: c.name for c in categories}

    cards = list(db.scalars(select(UserCard).where(UserCard.user_id == user_id)).all())
    card_id_map = {c.id: c.name for c in cards}

    type_display = {"income": "수입", "expense": "지출", "transfer": "이체"}

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "거래내역"

    headers = ["날짜", "유형", "금액", "내역", "카테고리", "결제수단", "카드명"]
    header_fill = PatternFill(start_color="3182F6", end_color="3182F6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    widths = [14, 8, 14, 30, 12, 12, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Data rows
    payment_display = {
        "cash": "현금", "credit_card": "신용카드",
        "debit_card": "체크카드", "bank": "계좌이체",
    }

    for tx in transactions:
        ws.append([
            tx.transacted_at.strftime("%Y-%m-%d") if tx.transacted_at else "",
            type_display.get(tx.type, tx.type),
            float(tx.amount),
            tx.description or "",
            cat_id_map.get(tx.category_id, "") if tx.category_id else "",
            payment_display.get(tx.payment_type, tx.payment_type or ""),
            card_id_map.get(tx.user_card_id, "") if tx.user_card_id else "",
        ])

    # Format amount column as number
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = "#,##0"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
