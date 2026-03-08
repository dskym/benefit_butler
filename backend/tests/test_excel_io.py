# backend/tests/test_excel_io.py
"""Tests for Excel import/export endpoints."""
from datetime import datetime, timezone
from io import BytesIO

import openpyxl
import xlwt
import pytest

from app.services import import_cache


def make_xlsx(headers: list[str], rows: list[list]) -> BytesIO:
    """Create an in-memory xlsx file from headers and rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def make_xls(headers: list[str], rows: list[list]) -> BytesIO:
    """Create an in-memory xls (Excel 97-2003) file from headers and rows."""
    wb = xlwt.Workbook()
    ws = wb.add_sheet("Sheet1")
    for col_idx, header in enumerate(headers):
        ws.write(0, col_idx, header)
    for row_idx, row in enumerate(rows):
        for col_idx, value in enumerate(row):
            ws.write(row_idx + 1, col_idx, value)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Fixtures ─────────────────────────────────────────────────────────────────

@pytest.fixture(autouse=True)
def _clear_import_cache():
    """Ensure import cache is clean for every test."""
    import_cache.clear()
    yield
    import_cache.clear()


def _create_card(client, auth_headers, name="테스트카드"):
    """Helper to create a user card and return its ID."""
    resp = client.post(
        "/api/v1/cards/",
        json={"type": "credit_card", "name": name},
        headers=auth_headers,
    )
    assert resp.status_code == 201
    return resp.json()["id"]


def _create_category(client, auth_headers, name="테스트카테고리", cat_type="expense"):
    """Helper to create a category and return its ID."""
    resp = client.post(
        "/api/v1/categories/",
        json={"name": name, "type": cat_type, "color": "#FF0000"},
        headers=auth_headers,
    )
    assert resp.status_code == 201
    return resp.json()["id"]


# ── Import Preview Tests ────────────────────────────────────────────────────


class TestImportPreview:
    def test_korean_headers_auto_detected(self, client, auth_headers):
        xlsx = make_xlsx(
            ["날짜", "금액", "내역", "구분", "카테고리"],
            [["2024-01-15", 15000, "스타벅스", "지출", "식비"]],
        )
        resp = client.post(
            "/api/v1/transactions/import/preview",
            files={"file": ("test.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["total_rows"] == 1
        assert data["auto_mapping"]["transacted_at"] == 0
        assert data["auto_mapping"]["amount"] == 1
        assert data["auto_mapping"]["description"] == 2
        assert data["auto_mapping"]["type"] == 3
        assert data["auto_mapping"]["category_name"] == 4

    def test_english_headers_auto_detected(self, client, auth_headers):
        xlsx = make_xlsx(
            ["date", "amount", "description", "type"],
            [["2024-01-15", 5000, "Coffee", "expense"]],
        )
        resp = client.post(
            "/api/v1/transactions/import/preview",
            files={"file": ("test.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["auto_mapping"]["transacted_at"] == 0
        assert data["auto_mapping"]["amount"] == 1
        assert data["auto_mapping"]["description"] == 2

    def test_preview_max_10_rows(self, client, auth_headers):
        rows = [[f"2024-01-{i+1:02d}", 1000 * (i + 1), f"Item {i}"] for i in range(20)]
        xlsx = make_xlsx(["날짜", "금액", "내역"], rows)
        resp = client.post(
            "/api/v1/transactions/import/preview",
            files={"file": ("test.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        data = resp.json()
        assert len(data["preview_rows"]) == 10
        assert data["total_rows"] == 20

    def test_xls_korean_headers_auto_detected(self, client, auth_headers):
        xls = make_xls(
            ["날짜", "금액", "내역", "구분"],
            [["2024-01-15", 15000, "스타벅스", "지출"]],
        )
        resp = client.post(
            "/api/v1/transactions/import/preview",
            files={"file": ("test.xls", xls, "application/vnd.ms-excel")},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["total_rows"] == 1
        assert data["auto_mapping"]["transacted_at"] == 0
        assert data["auto_mapping"]["amount"] == 1
        assert data["auto_mapping"]["description"] == 2

    def test_xls_confirm_creates_transactions(self, client, auth_headers):
        xls = make_xls(
            ["날짜", "금액", "내역"],
            [["2024-01-15", 15000, "스타벅스"]],
        )
        resp = client.post(
            "/api/v1/transactions/import/preview",
            files={"file": ("test.xls", xls, "application/vnd.ms-excel")},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        import_id = resp.json()["import_id"]

        resp = client.post(
            "/api/v1/transactions/import/confirm",
            json={
                "import_id": import_id,
                "mapping": {"transacted_at": 0, "amount": 1, "description": 2},
            },
            headers=auth_headers,
        )
        assert resp.status_code == 201
        assert resp.json()["created_count"] == 1

    def test_reject_non_xlsx(self, client, auth_headers):
        buf = BytesIO(b"not an xlsx file")
        resp = client.post(
            "/api/v1/transactions/import/preview",
            files={"file": ("test.csv", buf, "text/csv")},
            headers=auth_headers,
        )
        assert resp.status_code == 400

    def test_reject_invalid_xlsx(self, client, auth_headers):
        buf = BytesIO(b"not xlsx content")
        resp = client.post(
            "/api/v1/transactions/import/preview",
            files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=auth_headers,
        )
        assert resp.status_code == 400

    def test_reject_empty_data(self, client, auth_headers):
        xlsx = make_xlsx(["날짜", "금액", "내역"], [])
        resp = client.post(
            "/api/v1/transactions/import/preview",
            files={"file": ("test.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=auth_headers,
        )
        assert resp.status_code == 400

    def test_requires_auth(self, client):
        xlsx = make_xlsx(["날짜", "금액"], [["2024-01-01", 1000]])
        resp = client.post(
            "/api/v1/transactions/import/preview",
            files={"file": ("test.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 403


# ── Import Confirm Tests ────────────────────────────────────────────────────


class TestImportConfirm:
    def _preview(self, client, auth_headers, headers, rows):
        """Upload a preview and return the import_id."""
        xlsx = make_xlsx(headers, rows)
        resp = client.post(
            "/api/v1/transactions/import/preview",
            files={"file": ("test.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        return resp.json()["import_id"]

    def test_create_transactions(self, client, auth_headers):
        import_id = self._preview(
            client, auth_headers,
            ["날짜", "금액", "내역"],
            [
                ["2024-01-15", 15000, "스타벅스"],
                ["2024-01-16", 25000, "편의점"],
            ],
        )
        resp = client.post(
            "/api/v1/transactions/import/confirm",
            json={
                "import_id": import_id,
                "mapping": {"transacted_at": 0, "amount": 1, "description": 2},
            },
            headers=auth_headers,
        )
        assert resp.status_code == 201
        data = resp.json()
        assert data["created_count"] == 2
        assert data["duplicate_count"] == 0
        assert data["error_count"] == 0

        # Verify transactions exist
        txs = client.get("/api/v1/transactions/", headers=auth_headers).json()
        assert len(txs) == 2

    def test_skip_duplicates(self, client, auth_headers):
        # Create an existing transaction
        client.post(
            "/api/v1/transactions/",
            json={
                "type": "expense",
                "amount": 15000,
                "description": "스타벅스",
                "transacted_at": "2024-01-15T00:00:00Z",
            },
            headers=auth_headers,
        )

        import_id = self._preview(
            client, auth_headers,
            ["날짜", "금액", "내역"],
            [
                ["2024-01-15", 15000, "스타벅스"],  # duplicate
                ["2024-01-16", 25000, "편의점"],     # new
            ],
        )
        resp = client.post(
            "/api/v1/transactions/import/confirm",
            json={
                "import_id": import_id,
                "mapping": {"transacted_at": 0, "amount": 1, "description": 2},
            },
            headers=auth_headers,
        )
        assert resp.status_code == 201
        data = resp.json()
        assert data["created_count"] == 1
        assert data["duplicate_count"] == 1

    def test_category_name_matching(self, client, auth_headers):
        _create_category(client, auth_headers, "식비", "expense")
        import_id = self._preview(
            client, auth_headers,
            ["날짜", "금액", "내역", "카테고리"],
            [["2024-01-15", 15000, "스타벅스", "식비"]],
        )
        resp = client.post(
            "/api/v1/transactions/import/confirm",
            json={
                "import_id": import_id,
                "mapping": {"transacted_at": 0, "amount": 1, "description": 2, "category_name": 3},
            },
            headers=auth_headers,
        )
        assert resp.status_code == 201
        assert resp.json()["created_count"] == 1

        txs = client.get("/api/v1/transactions/", headers=auth_headers).json()
        assert txs[0]["category_id"] is not None

    def test_card_name_matching(self, client, auth_headers):
        _create_card(client, auth_headers, "신한카드")
        import_id = self._preview(
            client, auth_headers,
            ["날짜", "금액", "내역", "카드명"],
            [["2024-01-15", 15000, "스타벅스", "신한카드"]],
        )
        resp = client.post(
            "/api/v1/transactions/import/confirm",
            json={
                "import_id": import_id,
                "mapping": {"transacted_at": 0, "amount": 1, "description": 2, "card_name": 3},
            },
            headers=auth_headers,
        )
        assert resp.status_code == 201
        assert resp.json()["created_count"] == 1

        txs = client.get("/api/v1/transactions/", headers=auth_headers).json()
        assert txs[0]["user_card_id"] is not None

    def test_expired_import_id_404(self, client, auth_headers):
        resp = client.post(
            "/api/v1/transactions/import/confirm",
            json={
                "import_id": "nonexistent-id",
                "mapping": {"transacted_at": 0, "amount": 1},
            },
            headers=auth_headers,
        )
        assert resp.status_code == 404

    def test_missing_required_mapping_400(self, client, auth_headers):
        import_id = self._preview(
            client, auth_headers,
            ["날짜", "금액"],
            [["2024-01-15", 1000]],
        )
        # Missing amount mapping
        resp = client.post(
            "/api/v1/transactions/import/confirm",
            json={
                "import_id": import_id,
                "mapping": {"transacted_at": 0},
            },
            headers=auth_headers,
        )
        assert resp.status_code == 400

    def test_various_date_formats(self, client, auth_headers):
        import_id = self._preview(
            client, auth_headers,
            ["날짜", "금액"],
            [
                ["2024-01-15", 1000],
                ["2024/02/20", 2000],
                ["2024.03.25", 3000],
            ],
        )
        resp = client.post(
            "/api/v1/transactions/import/confirm",
            json={
                "import_id": import_id,
                "mapping": {"transacted_at": 0, "amount": 1},
            },
            headers=auth_headers,
        )
        assert resp.status_code == 201
        assert resp.json()["created_count"] == 3

    def test_korean_type_mapping(self, client, auth_headers):
        import_id = self._preview(
            client, auth_headers,
            ["날짜", "금액", "구분"],
            [
                ["2024-01-15", 1000, "지출"],
                ["2024-01-16", 2000, "수입"],
                ["2024-01-17", 3000, "이체"],
            ],
        )
        resp = client.post(
            "/api/v1/transactions/import/confirm",
            json={
                "import_id": import_id,
                "mapping": {"transacted_at": 0, "amount": 1, "type": 2},
            },
            headers=auth_headers,
        )
        assert resp.status_code == 201
        assert resp.json()["created_count"] == 3

        txs = client.get("/api/v1/transactions/", headers=auth_headers).json()
        types = {tx["type"] for tx in txs}
        assert types == {"expense", "income", "transfer"}

    def test_amount_with_commas_and_currency(self, client, auth_headers):
        import_id = self._preview(
            client, auth_headers,
            ["날짜", "금액"],
            [
                ["2024-01-15", "15,000원"],
                ["2024-01-16", "₩25,000"],
            ],
        )
        resp = client.post(
            "/api/v1/transactions/import/confirm",
            json={
                "import_id": import_id,
                "mapping": {"transacted_at": 0, "amount": 1},
            },
            headers=auth_headers,
        )
        assert resp.status_code == 201
        assert resp.json()["created_count"] == 2

    def test_requires_auth(self, client):
        resp = client.post(
            "/api/v1/transactions/import/confirm",
            json={
                "import_id": "some-id",
                "mapping": {"transacted_at": 0, "amount": 1},
            },
        )
        assert resp.status_code == 403


# ── Export Tests ─────────────────────────────────────────────────────────────


class TestExport:
    def test_export_month(self, client, auth_headers):
        # Create a transaction
        client.post(
            "/api/v1/transactions/",
            json={
                "type": "expense",
                "amount": 15000,
                "description": "스타벅스",
                "transacted_at": "2024-01-15T00:00:00Z",
            },
            headers=auth_headers,
        )

        resp = client.get(
            "/api/v1/transactions/export?period=month&year=2024&month=1",
            headers=auth_headers,
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]
        assert "transactions_2024_01" in resp.headers["content-disposition"]

        # Parse the xlsx
        wb = openpyxl.load_workbook(BytesIO(resp.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0] == ("날짜", "유형", "금액", "내역", "카테고리", "결제수단", "카드명")
        assert len(rows) == 2  # header + 1 data row
        assert rows[1][3] == "스타벅스"

    def test_export_year(self, client, auth_headers):
        client.post(
            "/api/v1/transactions/",
            json={
                "type": "income",
                "amount": 3000000,
                "description": "급여",
                "transacted_at": "2024-06-25T00:00:00Z",
            },
            headers=auth_headers,
        )
        resp = client.get(
            "/api/v1/transactions/export?period=year&year=2024",
            headers=auth_headers,
        )
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(BytesIO(resp.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 2

    def test_export_all(self, client, auth_headers):
        resp = client.get(
            "/api/v1/transactions/export?period=all",
            headers=auth_headers,
        )
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(BytesIO(resp.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 1  # header only

    def test_export_empty_data(self, client, auth_headers):
        resp = client.get(
            "/api/v1/transactions/export?period=month&year=2024&month=1",
            headers=auth_headers,
        )
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(BytesIO(resp.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 1  # header only

    def test_export_includes_category_card_names(self, client, auth_headers):
        cat_id = _create_category(client, auth_headers, "식비", "expense")
        card_id = _create_card(client, auth_headers, "신한카드")
        client.post(
            "/api/v1/transactions/",
            json={
                "type": "expense",
                "amount": 15000,
                "description": "점심",
                "transacted_at": "2024-01-15T00:00:00Z",
                "category_id": cat_id,
                "user_card_id": card_id,
                "payment_type": "credit_card",
            },
            headers=auth_headers,
        )
        resp = client.get(
            "/api/v1/transactions/export?period=month&year=2024&month=1",
            headers=auth_headers,
        )
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(BytesIO(resp.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 2
        data_row = rows[1]
        assert "식비" in str(data_row)
        assert "신한카드" in str(data_row)
        assert "신용카드" in str(data_row)

    def test_requires_auth(self, client):
        resp = client.get("/api/v1/transactions/export?period=all")
        assert resp.status_code == 403
